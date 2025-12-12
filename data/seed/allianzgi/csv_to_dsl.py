#!/usr/bin/env python3
"""
AllianzGI Fund Data to DSL Converter
============================================================================
Converts downloaded AllianzGI fund XLSX/CSV exports to DSL load commands.

Usage:
    python csv_to_dsl.py out/LU__AGI_LUX__FundList_*.xlsx > lu_funds.dsl

Expected columns (based on AllianzGI "Daily Fund Price" export):
    - Fund (sub-fund name)
    - Isin
    - As of (NAV date)
    - Asset Class
    - Share Class Currency
    - Share Class (designation like "A (EUR)", "IT (USD)")
    - NAV
============================================================================
"""

import csv
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional

try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@dataclass
class ShareClass:
    """Parsed share class from CSV"""

    fund_name: str
    share_class_name: str
    isin: str
    currency: str
    asset_class: str
    share_class_type: str  # RETAIL, INSTITUTIONAL
    distribution_type: str  # ACC, DIST
    is_hedged: bool
    sfdr_category: str


def parse_share_class_type(name: str) -> tuple[str, str, bool]:
    """
    Parse share class designation to extract type, dist, and hedge info.
    Examples:
        "A - EUR" -> RETAIL, ACC, False
        "IT - USD" -> INSTITUTIONAL, ACC, False
        "ADM - EUR" -> RETAIL, DIST, False
        "AT (H2-EUR) - EUR" -> RETAIL, ACC, True
    """
    name_upper = name.upper()

    # Determine share class type
    if any(x in name_upper for x in [" IT ", " WT ", " I ", " W ", " PT "]):
        share_type = "INSTITUTIONAL"
    else:
        share_type = "RETAIL"

    # Determine distribution type
    if any(x in name_upper for x in ["DM ", "ADM", " D ", " DIS"]):
        dist_type = "DIST"
    else:
        dist_type = "ACC"

    # Determine if hedged
    is_hedged = "H2" in name_upper or "(H-" in name_upper or "HEDGED" in name_upper

    return share_type, dist_type, is_hedged


def extract_subfund_name(full_name: str) -> str:
    """
    Extract sub-fund name from full share class name.
    "Allianz Global Artificial Intelligence - A - EUR" -> "Allianz Global Artificial Intelligence"
    """
    # Split on " - " and take everything before the share class designation
    parts = full_name.split(" - ")
    if len(parts) >= 2:
        return parts[0].strip()
    return full_name


def sanitize_var_name(name: str) -> str:
    """Convert name to valid DSL variable name"""
    # Remove special chars, convert to lowercase with underscores
    clean = re.sub(r"[^a-zA-Z0-9]+", "_", name.lower())
    clean = re.sub(r"_+", "_", clean).strip("_")
    return clean[:50]  # Truncate if too long


def parse_xlsx(filepath: Path, jurisdiction: str, manco_code: str) -> List[ShareClass]:
    """Parse AllianzGI XLSX export (Daily Fund Price format)"""
    if not HAS_OPENPYXL:
        print(
            "Error: openpyxl required for XLSX files. Install with: pip install openpyxl",
            file=sys.stderr,
        )
        sys.exit(1)

    share_classes = []
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    # Find header row (contains "Fund", "Isin", etc.)
    header_row = None
    headers = {}
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=20, values_only=True), 1
    ):
        row_values = [str(v).lower() if v else "" for v in row]
        if "fund" in row_values and "isin" in row_values:
            header_row = row_idx
            for col_idx, val in enumerate(row):
                if val:
                    headers[str(val).lower().strip()] = col_idx
            break

    if header_row is None:
        print(f"Error: Could not find header row in {filepath}", file=sys.stderr)
        return []

    # Map column indices
    fund_col = headers.get("fund", 0)
    isin_col = headers.get("isin", 1)
    asset_class_col = headers.get("asset class", 3)
    currency_col = headers.get("share class currency", 4)
    share_class_col = headers.get("share class", 5)

    # Parse data rows
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        try:
            fund_name = str(row[fund_col]).strip() if row[fund_col] else ""
            isin = str(row[isin_col]).strip() if row[isin_col] else ""
            share_class_designation = (
                str(row[share_class_col]).strip() if row[share_class_col] else ""
            )
            currency = (
                str(row[currency_col]).strip()[:3] if row[currency_col] else "EUR"
            )
            asset_class = (
                str(row[asset_class_col]).strip() if row[asset_class_col] else "EQUITY"
            )

            if not isin or not fund_name:
                continue

            # Combine fund name and share class designation
            share_class_name = (
                f"{fund_name} - {share_class_designation}"
                if share_class_designation
                else fund_name
            )

            share_type, dist_type, is_hedged = parse_share_class_type(
                share_class_designation or share_class_name
            )

            sc = ShareClass(
                fund_name=fund_name,
                share_class_name=share_class_name,
                isin=isin,
                currency=currency,
                asset_class=asset_class,
                share_class_type=share_type,
                distribution_type=dist_type,
                is_hedged=is_hedged,
                sfdr_category="",
            )
            share_classes.append(sc)

        except Exception as e:
            print(f"# Warning: Failed to parse row: {e}", file=sys.stderr)
            continue

    wb.close()
    return share_classes


def parse_csv(filepath: Path, jurisdiction: str, manco_code: str) -> List[ShareClass]:
    """Parse AllianzGI CSV export"""
    share_classes = []

    with open(filepath, "r", encoding="utf-8-sig") as f:
        # Try to detect delimiter
        sample = f.read(2000)
        f.seek(0)

        if "\t" in sample:
            delimiter = "\t"
        elif ";" in sample:
            delimiter = ";"
        else:
            delimiter = ","

        reader = csv.DictReader(f, delimiter=delimiter)

        # Normalize column names (handle different languages)
        col_map = {}
        for col in reader.fieldnames or []:
            col_lower = col.lower().strip()
            if "fund" in col_lower or "fonds" in col_lower:
                col_map["fund_name"] = col
            elif col_lower == "isin":
                col_map["isin"] = col
            elif "share" in col_lower or "anteils" in col_lower or "class" in col_lower:
                col_map["share_class"] = col
            elif "currency" in col_lower or "währung" in col_lower:
                col_map["currency"] = col
            elif "asset" in col_lower:
                col_map["asset_class"] = col
            elif "sfdr" in col_lower:
                col_map["sfdr"] = col

        for row in reader:
            try:
                # Get share class name - might be separate or part of fund name
                full_name = row.get(col_map.get("fund_name", ""), "")
                share_class_suffix = row.get(col_map.get("share_class", ""), "")

                if share_class_suffix and share_class_suffix not in full_name:
                    share_class_name = f"{full_name} - {share_class_suffix}"
                else:
                    share_class_name = full_name

                isin = row.get(col_map.get("isin", ""), "").strip()
                currency = row.get(col_map.get("currency", ""), "EUR").strip()[:3]
                asset_class = row.get(col_map.get("asset_class", ""), "EQUITY").strip()
                sfdr = row.get(col_map.get("sfdr", ""), "").strip()

                if not isin or not full_name:
                    continue

                share_type, dist_type, is_hedged = parse_share_class_type(
                    share_class_name
                )

                sc = ShareClass(
                    fund_name=extract_subfund_name(full_name),
                    share_class_name=share_class_name,
                    isin=isin,
                    currency=currency,
                    asset_class=asset_class,
                    share_class_type=share_type,
                    distribution_type=dist_type,
                    is_hedged=is_hedged,
                    sfdr_category=sfdr,
                )
                share_classes.append(sc)

            except Exception as e:
                print(f"# Warning: Failed to parse row: {e}", file=sys.stderr)
                continue

    return share_classes


def parse_file(filepath: Path, jurisdiction: str, manco_code: str) -> List[ShareClass]:
    """Parse fund data file (XLSX or CSV)"""
    suffix = filepath.suffix.lower()
    if suffix == ".xlsx":
        return parse_xlsx(filepath, jurisdiction, manco_code)
    else:
        return parse_csv(filepath, jurisdiction, manco_code)


def generate_dsl(
    share_classes: List[ShareClass],
    jurisdiction: str,
    manco_code: str,
    umbrella_name: str = "Allianz Global Investors Fund",
    cbu_name: str = "Allianz Global Investors Group",
) -> str:
    """Generate DSL commands from parsed share classes (S-expression syntax)"""

    # Group by sub-fund
    subfunds: Dict[str, List[ShareClass]] = {}
    for sc in share_classes:
        if sc.fund_name not in subfunds:
            subfunds[sc.fund_name] = []
        subfunds[sc.fund_name].append(sc)

    lines = []
    lines.append(f";; AllianzGI {jurisdiction} Fund Load - Full Export")
    lines.append(f";; Generated from regulatory XLSX export")
    lines.append(f";; ManCo: {manco_code}")
    lines.append(f";; Sub-funds: {len(subfunds)}")
    lines.append(f";; Share classes: {len(share_classes)}")
    lines.append(f";; " + "=" * 70)
    lines.append("")
    lines.append(f";; PREREQUISITE: Run 03_load_allianzgi.dsl first to create:")
    lines.append(f";;   CBU: {cbu_name}")
    lines.append(f";;   Umbrella: {umbrella_name}")
    lines.append(f";; " + "=" * 70)
    lines.append("")

    # Generate sub-fund and share class commands
    for subfund_name, scs in subfunds.items():
        var_name = sanitize_var_name(subfund_name)

        # Determine base currency from most common share class currency
        currencies = [sc.currency for sc in scs if not sc.is_hedged]
        base_currency = (
            max(set(currencies), key=currencies.count) if currencies else "EUR"
        )

        lines.append(f";; Sub-fund: {subfund_name} ({len(scs)} share classes)")
        lines.append(
            f'(fund.create-subfund :name "{subfund_name}" '
            f':umbrella-id "{umbrella_name}" '
            f':base-currency "{base_currency}" '
            f':cbu-id "{cbu_name}" '
            f":as @sf_{var_name})"
        )

        # Generate share classes
        for sc in scs:
            sc_var = sanitize_var_name(sc.isin)
            hedged_arg = " :hedged true" if sc.is_hedged else ""
            lines.append(
                f'(fund.create-share-class :name "{sc.share_class_name}" '
                f":subfund-id @sf_{var_name} "
                f':share-class-type "{sc.share_class_type}" '
                f':distribution-type "{sc.distribution_type}" '
                f':currency "{sc.currency}"{hedged_arg} '
                f':isin "{sc.isin}" '
                f":as @sc_{sc_var})"
            )

        lines.append("")

    # Summary
    lines.append(f";; " + "=" * 70)
    lines.append(
        f";; SUMMARY: {len(subfunds)} sub-funds, {len(share_classes)} share classes"
    )
    lines.append(f";; " + "=" * 70)

    return "\n".join(lines)


def main():
    if len(sys.argv) < 2:
        print(
            "Usage: python csv_to_dsl.py <csv_file> [jurisdiction] [manco_code]",
            file=sys.stderr,
        )
        print(
            "Example: python csv_to_dsl.py out/LU__AGI_LUX__funds.csv LU AGI_LUX",
            file=sys.stderr,
        )
        sys.exit(1)

    filepath = Path(sys.argv[1])

    # Try to extract jurisdiction and manco from filename
    filename = filepath.stem
    parts = filename.split("__")

    jurisdiction = (
        sys.argv[2] if len(sys.argv) > 2 else (parts[0] if len(parts) >= 1 else "LU")
    )
    manco_code = (
        sys.argv[3]
        if len(sys.argv) > 3
        else (parts[1] if len(parts) >= 2 else "AGI_LUX")
    )

    share_classes = parse_file(filepath, jurisdiction, manco_code)
    dsl = generate_dsl(share_classes, jurisdiction, manco_code)
    print(dsl)


if __name__ == "__main__":
    main()
